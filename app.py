"""
ICode 签到管理系统 v3.0
Design: Code Planet — deep space aesthetic, electric blue + neon orange
"""
from flask import Flask, render_template, request, jsonify, session, redirect, send_file
import hashlib, json, io, os
from datetime import datetime
from functools import wraps
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pymysql
import pymysql.cursors

app = Flask(__name__)
import os
app.secret_key = os.environ.get('SECRET_KEY', 'icode-planet-v3-2025-secret')

# ─── MySQL 连接配置 ────────────────────────────────────────────────
DB_CONFIG = {
    'host':     'localhost',
    'port':     3306,
    'user':     'icode',
    'password': 'Icode2025!',
    'db':       'icode_checkin',
    'charset':  'utf8mb4',
    'cursorclass': pymysql.cursors.DictCursor,
    'autocommit': False,
}

# ─── helpers ───────────────────────────────────────────────────────
def db():
    return pymysql.connect(**DB_CONFIG)

def sha(s):
    return hashlib.sha256(s.encode()).hexdigest()

def now_str():
    return datetime.now().strftime('%Y-%m-%d %H:%M:%S')

def db_query(sql, params=(), one=False):
    """执行查询，返回字典列表或单行字典"""
    sql = sql.replace('?', '%s')
    conn = db()
    try:
        c = conn.cursor()
        c.execute(sql, params)
        return c.fetchone() if one else c.fetchall()
    finally:
        conn.close()

def db_exec(sql, params=()):
    """执行写操作，返回 lastrowid"""
    sql = sql.replace('?', '%s')
    conn = db()
    try:
        c = conn.cursor()
        c.execute(sql, params)
        conn.commit()
        return c.lastrowid
    finally:
        conn.close()

def db_exec_many(sqls_params):
    """在同一个事务里执行多条写操作"""
    conn = db()
    try:
        c = conn.cursor()
        for sql, params in sqls_params:
            c.execute(sql.replace('?', '%s'), params)
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

def init_db():
    """MySQL版：补充新字段，确保主管理员存在"""
    conn = db(); c = conn.cursor()
    # players 表补新字段
    for col, coldef in [
        ('package', 'VARCHAR(100) DEFAULT ""'),
        ('extra_data', 'TEXT'),
    ]:
        try:
            c.execute(f"ALTER TABLE players ADD COLUMN {col} {coldef}")
            print(f"  ✅ players.{col} 已添加")
        except Exception:
            pass
    # competitions 表补新字段
    try:
        c.execute("ALTER TABLE competitions ADD COLUMN extra_fields TEXT")
        print("  ✅ competitions.extra_fields 已添加")
    except Exception:
        pass
    # checkin_logs 补签到人字段
    for col, coldef in [
        ('operator_name', "VARCHAR(100) DEFAULT ''"),
        ('operator_phone', "VARCHAR(50) DEFAULT ''"),
    ]:
        try:
            c.execute(f"ALTER TABLE checkin_logs ADD COLUMN {col} {coldef}")
            print(f"  ✅ checkin_logs.{col} 已添加")
        except Exception:
            pass
    # players 补签到人字段
    for col, coldef in [
        ('checkin_operator_name', "VARCHAR(100) DEFAULT ''"),
        ('checkin_operator_phone', "VARCHAR(50) DEFAULT ''"),
    ]:
        try:
            c.execute(f"ALTER TABLE players ADD COLUMN {col} {coldef}")
            print(f"  ✅ players.{col} 已添加")
        except Exception:
            pass
    conn.commit(); conn.close()
    existing = db_query("SELECT id FROM admins WHERE is_main=1", one=True)
    if not existing:
        db_exec("INSERT INTO admins(name,phone,password,is_main,permissions) VALUES(%s,%s,%s,1,%s)",
                ('主管理员', 'admin', sha('admin123'), '{"all":true}'))
        print("  ✅ 已创建默认主管理员 admin/admin123")

def admin_required(f):
    @wraps(f)
    def w(*a, **k):
        if 'admin_id' not in session:
            return jsonify({'error': '未授权'}), 401
        return f(*a, **k)
    return w

def get_me():
    if 'admin_id' not in session: return None
    a = db_query("SELECT * FROM admins WHERE id=?", (session['admin_id'],), one=True)
    if not a:
        session.clear(); return None
    if not a['is_main'] and not a.get('is_active', 1):
        session.clear(); return None  # 禁用账号强制退出
    return a

def can(admin, perm):
    if not admin: return False
    if admin['is_main']: return True
    p = json.loads(admin['permissions'] or '{}')
    return p.get('all') or p.get(perm, False)

def comp_perm(admin, comp_id):
    """返回该管理员对指定赛事的权限级别：'edit' / 'view' / None"""
    if not admin: return None
    if admin['is_main']: return 'edit'
    comp = db_query("SELECT created_by, comp_admins FROM competitions WHERE id=?", (comp_id,), one=True)
    if not comp: return None
    if comp['created_by'] == admin['id']: return 'edit'
    try:
        for ca in json.loads(comp['comp_admins'] or '[]'):
            if ca.get('admin_id') == admin['id']:
                return ca.get('perm', 'view')  # 'edit' or 'view'
    except Exception:
        pass
    return None

def admin_owns_comp(admin, comp_id):
    """兼容旧调用：有任意权限即返回 True"""
    return comp_perm(admin, comp_id) is not None

def can_view_comp(admin, comp_id):
    return comp_perm(admin, comp_id) in ('view', 'edit')

def can_edit_comp(admin, comp_id):
    return comp_perm(admin, comp_id) == 'edit' 

# ═══════════════════════════════════════════════════════════════════
# PUBLIC PLAYER ROUTES
# ═══════════════════════════════════════════════════════════════════

@app.route('/favicon.ico')
def favicon():
    return '', 204

@app.route('/')
def player_root():
    return render_template('player.html')

@app.route('/c/<int:cid>')
def player_comp(cid):
    return render_template('player.html', preset_cid=cid)

@app.route('/api/pub/competition/<int:cid>')
def pub_competition(cid):
    c = db_query(
        "SELECT id,name,description,description_images,album_url,banner_text,banner_color,banner_accent,`groups`,"
        "display_fields,query_field,query_hint,location,start_time,end_time,extra_fields "
        "FROM competitions WHERE id=? AND is_active=1", (cid,), one=True)
    if not c: return jsonify({'error': '该赛事不存在或已下线'}), 404
    return jsonify(dict(c))

@app.route('/api/pub/competitions')
def pub_competitions():
    rows = db_query(
        "SELECT id,name,description,banner_text,banner_color,banner_accent,location,start_time "
        "FROM competitions WHERE is_active=1 ORDER BY id DESC")
    return jsonify([dict(r) for r in rows])

@app.route('/api/pub/query', methods=['POST'])
def pub_query():
    d = request.json or {}
    cid = d.get('competition_id')
    q = d.get('query', '').strip()
    if not q: return jsonify({'error': '请输入查询内容'}), 400
    if not cid: return jsonify({'error': '比赛参数缺失'}), 400
    comp = db_query("SELECT display_fields,query_field,extra_fields FROM competitions WHERE id=?", (cid,), one=True)
    if not comp: return jsonify({'error': '赛事不存在'}), 404
    display_fields = json.loads(comp['display_fields'])
    extra_field_defs = json.loads(comp['extra_fields'] or '[]')
    query_fields = [f.strip() for f in (comp['query_field'] or 'player_no,account').split(',')]
    tokens = [t for t in q.split() if t]
    results = []
    seen = set()
    for t in tokens:
        conds = ' OR '.join([f"{f}=%s" for f in query_fields])
        row = db_query(
            f"SELECT * FROM players WHERE competition_id=%s AND ({conds})",
            (cid, *([t] * len(query_fields))), one=True)
        if row and row['id'] not in seen:
            results.append(dict(row))
            seen.add(row['id'])
    if not results: return jsonify({'error': '未找到选手，请检查编号是否正确'}), 404
    return jsonify({'players': results, 'display_fields': display_fields, 'extra_field_defs': extra_field_defs})

@app.route('/api/pub/checkin', methods=['POST'])
def pub_checkin():
    d = request.json or {}
    ids = d.get('player_ids', [])
    if not ids: return jsonify({'error': '请选择要签到的选手'}), 400
    operator_name  = str(d.get('operator_name', '选手自助')).strip()[:100]
    operator_phone = str(d.get('operator_phone', '')).strip()[:50]
    nw = now_str()
    names = []
    ops = []
    cid_from_body = d.get('competition_id')
    for pid in ids:
        p = db_query("SELECT * FROM players WHERE id=?", (pid,), one=True)
        if not p or p['checked_in']: continue
        comp_check = db_query("SELECT is_active FROM competitions WHERE id=?", (p['competition_id'],), one=True)
        if not comp_check or not comp_check['is_active']: continue
        if cid_from_body and str(p['competition_id']) != str(cid_from_body): continue
        ops.append(("UPDATE players SET checked_in=1,checkin_time=?,checkin_operator_name=?,checkin_operator_phone=? WHERE id=?",
                    (nw, operator_name, operator_phone, pid)))
        ops.append(("INSERT INTO checkin_logs(player_id,competition_id,checkin_time,operator_name,operator_phone) VALUES(?,?,?,?,?)",
                    (pid, p['competition_id'], nw, operator_name, operator_phone)))
        names.append(p['name'])
    if ops: db_exec_many(ops)
    return jsonify({'success': True, 'names': names})

# ═══════════════════════════════════════════════════════════════════
# ADMIN AUTH
# ═══════════════════════════════════════════════════════════════════

@app.route('/admin')
def admin_login_page():
    if 'admin_id' in session: return redirect('/admin/dashboard')
    return render_template('admin_login.html')

@app.route('/admin/dashboard')
def admin_dash():
    if 'admin_id' not in session: return redirect('/admin')
    from flask import make_response
    resp = make_response(render_template('admin_dashboard.html'))
    resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    resp.headers['Pragma'] = 'no-cache'
    resp.headers['Expires'] = '0'
    return resp

@app.route('/api/admin/login', methods=['POST'])
def admin_login():
    d = request.json or {}
    # 简单暴力破解防护：同一会话连续失败5次锁定10分钟
    fail_key = f"fail_{d.get('phone','')}"
    fail_info = session.get(fail_key, {'count':0,'until':0})
    import time
    if fail_info['until'] > time.time():
        return jsonify({'error': f'登录失败次数过多，请{int(fail_info["until"]-time.time()//60)+1}分钟后再试'}), 429
    a = db_query("SELECT * FROM admins WHERE phone=?", (d.get('phone', ''),), one=True)
    if not a or a['password'] != sha(d.get('password', '')):
        fail_info['count'] = fail_info.get('count',0) + 1
        if fail_info['count'] >= 5:
            fail_info['until'] = time.time() + 600  # 锁定10分钟
            fail_info['count'] = 0
        session[fail_key] = fail_info
        return jsonify({'error': '账号或密码错误'}), 401
    session.pop(fail_key, None)  # 登录成功清除失败计数
    if not a['is_main'] and not a.get('is_active', 1):
        return jsonify({'error': '账号已被禁用，请联系主管理员'}), 403
    session['admin_id'] = a['id']
    session['admin_name'] = a['name']
    session['is_main'] = a['is_main']
    return jsonify({'success': True, 'name': a['name'], 'is_main': a['is_main']})

@app.route('/api/admin/logout', methods=['POST'])
def admin_logout():
    session.clear()
    return jsonify({'success': True})

@app.route('/api/admin/me')
@admin_required
def admin_me():
    a = get_me()
    if not a: return jsonify({'error': '未授权'}), 401
    return jsonify({'id': a['id'], 'name': a['name'], 'phone': a['phone'],
                    'is_main': a['is_main'],
                    'permissions': json.loads(a['permissions'] or '{}')})

@app.route('/api/admin/change-password', methods=['POST'])
@admin_required
def change_pwd():
    d = request.json or {}
    a = db_query("SELECT * FROM admins WHERE id=?", (session['admin_id'],), one=True)
    if not a: return jsonify({'error': '用户不存在'}), 404
    if a['password'] != sha(d.get('old_password', '')):
        return jsonify({'error': '旧密码错误'}), 400
    if len(d.get('new_password','')) < 6:
        return jsonify({'error': '新密码至少6位'}), 400
    db_exec("UPDATE admins SET password=? WHERE id=?",
            (sha(d['new_password']), session['admin_id']))
    return jsonify({'success': True})

# ═══════════════════════════════════════════════════════════════════
# COMPETITIONS
# ═══════════════════════════════════════════════════════════════════

@app.route('/api/competitions', methods=['GET'])
@admin_required
def list_competitions():
    a = get_me()
    if not a: return jsonify({'error': '未授权'}), 401
    if a['is_main']:
        rows = db_query("""
            SELECT c.id,c.name,c.location,c.start_time,c.end_time,c.description,
            c.description_images,c.album_url,c.manager_name,c.comp_admins,
            c.banner_text,c.banner_color,c.banner_accent,c.`groups`,c.display_fields,
            c.query_field,c.query_hint,c.is_active,c.created_by,c.created_at,c.extra_fields,
            (SELECT COUNT(*) FROM players WHERE competition_id=c.id) pc,
            adm.name creator_name FROM competitions c
            LEFT JOIN admins adm ON c.created_by=adm.id ORDER BY c.id DESC""")
    else:
        # 包含自己创建的 + 被分配权限的赛事
        all_comps = db_query("""
            SELECT c.id,c.name,c.location,c.start_time,c.end_time,c.description,
            c.description_images,c.album_url,c.manager_name,c.comp_admins,
            c.banner_text,c.banner_color,c.banner_accent,c.`groups`,c.display_fields,
            c.query_field,c.query_hint,c.is_active,c.created_by,c.created_at,c.extra_fields,
            (SELECT COUNT(*) FROM players WHERE competition_id=c.id) pc,
            adm.name creator_name FROM competitions c
            LEFT JOIN admins adm ON c.created_by=adm.id ORDER BY c.id DESC""")
        rows = []
        for c in all_comps:
            if c['created_by'] == a['id']:
                rows.append(c)
            else:
                try:
                    for ca in json.loads(c['comp_admins'] or '[]'):
                        if ca.get('admin_id') == a['id']:
                            rows.append(c); break
                except Exception:
                    pass
    return jsonify([dict(r) for r in rows])

@app.route('/api/competitions', methods=['POST'])
@admin_required
def create_competition():
    a = get_me()
    if not a: return jsonify({'error': '未授权'}), 401
    if not can(a, 'add_competition'): return jsonify({'error': '无权限'}), 403
    d = request.json or {}
    if not d.get('name'): return jsonify({'error': '请填写赛事名称'}), 400
    db_exec("""INSERT INTO competitions(name,location,start_time,end_time,description,
                    description_images,album_url,manager_name,comp_admins,
                    banner_text,banner_color,banner_accent,`groups`,display_fields,
                    query_field,query_hint,is_active,created_by,extra_fields) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                 (d['name'], d.get('location',''), d.get('start_time',''), d.get('end_time',''),
                  d.get('description',''),
                  json.dumps(d.get('description_images',[]), ensure_ascii=False),
                  d.get('album_url',''), d.get('manager_name',''),
                  json.dumps(d.get('comp_admins',[]), ensure_ascii=False),
                  d.get('banner_text','欢迎参加ICode比赛'),
                  d.get('banner_color','#1a6fa8'), d.get('banner_accent','#0099cc'),
                  json.dumps(d.get('groups',[]), ensure_ascii=False),
                  json.dumps(d.get('display_fields',['name','school','group_name','session','seat_no','shirt_size','package']), ensure_ascii=False),
                  d.get('query_field','player_no,account'),
                  d.get('query_hint','请输入报名编号或选手账号'),
                  d.get('is_active',1), session['admin_id'],
                  json.dumps(d.get('extra_fields',[]), ensure_ascii=False)))
    return jsonify({'success': True})

@app.route('/api/competitions/<int:cid>', methods=['GET'])
@admin_required
def get_competition(cid):
    a = get_me()
    if not a: return jsonify({'error': '未授权'}), 401
    r = db_query("SELECT * FROM competitions WHERE id=?", (cid,), one=True)
    if not r: return jsonify({'error': '不存在'}), 404
    if not a['is_main'] and not can_view_comp(a, cid): return jsonify({'error': '无权限'}), 403
    return jsonify(dict(r))

@app.route('/api/competitions/<int:cid>', methods=['PUT'])
@admin_required
def update_competition(cid):
    a = get_me()
    if not a: return jsonify({'error': '未授权'}), 401
    if not can_edit_comp(a, cid): return jsonify({'error': '无权限，需要编辑权限'}), 403
    d = request.json or {}
    db_exec("""UPDATE competitions SET name=?,location=?,start_time=?,end_time=?,
                    description=?,description_images=?,album_url=?,manager_name=?,comp_admins=?,
                    banner_text=?,banner_color=?,banner_accent=?,`groups`=?,
                    display_fields=?,query_field=?,query_hint=?,is_active=?,extra_fields=? WHERE id=?""",
                 (d.get('name'), d.get('location',''), d.get('start_time',''), d.get('end_time',''),
                  d.get('description',''),
                  json.dumps(d.get('description_images',[]), ensure_ascii=False),
                  d.get('album_url',''), d.get('manager_name',''),
                  json.dumps(d.get('comp_admins',[]), ensure_ascii=False),
                  d.get('banner_text',''), d.get('banner_color','#1a6fa8'), d.get('banner_accent','#0099cc'),
                  json.dumps(d.get('groups',[]), ensure_ascii=False),
                  json.dumps(d.get('display_fields',[]), ensure_ascii=False),
                  d.get('query_field','player_no,account'),
                  d.get('query_hint',''), d.get('is_active',1),
                  json.dumps(d.get('extra_fields',[]), ensure_ascii=False), cid))
    return jsonify({'success': True})

@app.route('/api/competitions/<int:cid>', methods=['DELETE'])
@admin_required
def delete_competition(cid):
    a = get_me()
    if not a: return jsonify({'error': '未授权'}), 401
    if not can_edit_comp(a, cid): return jsonify({'error': '无权限，需要编辑权限'}), 403
    db_exec_many([
        ("DELETE FROM checkin_logs WHERE competition_id=?", (cid,)),
        ("DELETE FROM players WHERE competition_id=?", (cid,)),
        ("DELETE FROM competitions WHERE id=?", (cid,)),
    ])
    return jsonify({'success': True})

# ── 功能5：批量导入赛事 ──────────────────────────────────────────
@app.route('/api/competitions/import', methods=['POST'])
@admin_required
def import_competitions():
    a = get_me()
    if not a: return jsonify({'error': '未授权'}), 401
    if not can(a, 'add_competition'): return jsonify({'error': '无权限'}), 403
    upload = request.files.get('file')
    if not upload: return jsonify({'error': '请上传文件'}), 400
    wb = openpyxl.load_workbook(upload, data_only=True)
    ws = wb.active
    hdrs = [str(c.value).strip().rstrip('*').strip() if c.value else '' for c in ws[1]]
    col_map = {
        '赛事名称':'name','地点':'location','开始时间':'start_time','结束时间':'end_time',
        '欢迎语':'banner_text','组别':'groups','赛事说明':'description','云相册链接':'album_url',
        '负责人':'manager_name','查询字段':'query_field','查询提示':'query_hint','是否上线':'is_active',
        '子管理员手机号':'sub_phones','子管理员权限':'sub_perms',
    }
    _conn = db(); _cur = _conn.cursor(); cnt = 0; warnings = []
    for row_i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row): continue
        d = {}
        for i, hdr_name in enumerate(hdrs):
            if hdr_name in col_map and i < len(row):
                d[col_map[hdr_name]] = str(row[i]).strip() if row[i] is not None else ''
        if not d.get('name'):
            warnings.append(f'第{row_i}行：赛事名称为空，已跳过'); continue
        grps = [g.strip() for g in d.get('groups','').split(',') if g.strip()]
        is_active = 0 if d.get('is_active','') in ('0','否','下线','no') else 1
        # 解析子管理员
        comp_admins = []
        phones = [p.strip() for p in d.get('sub_phones','').split(',') if p.strip()]
        perms  = [p.strip() for p in d.get('sub_perms','').split(',') if p.strip()]
        for idx, phone in enumerate(phones):
            _cur.execute("SELECT id FROM admins WHERE phone=%s", (phone,)); adm = _cur.fetchone()
            if adm:
                perm = perms[idx] if idx < len(perms) else 'view'
                perm = 'edit' if perm == 'edit' else 'view'
                comp_admins.append({'admin_id': adm['id'], 'perm': perm})
            else:
                warnings.append(f'第{row_i}行：手机号 {phone} 未找到对应管理员，已忽略')
        _cur.execute("""INSERT INTO competitions
            (name,location,start_time,end_time,description,album_url,manager_name,comp_admins,
             banner_text,banner_color,banner_accent,`groups`,display_fields,
             query_field,query_hint,is_active,created_by)
            VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
            (d.get('name',''), d.get('location',''), d.get('start_time',''), d.get('end_time',''),
             d.get('description',''), d.get('album_url',''), d.get('manager_name',''),
             json.dumps(comp_admins, ensure_ascii=False),
             d.get('banner_text','欢迎参加ICode比赛'), '#1a6fa8', '#0099cc',
             json.dumps(grps, ensure_ascii=False),
             '["name","school","group_name","session","seat_no","shirt_size","package"]',
             d.get('query_field','player_no,account'),
             d.get('query_hint','请输入报名编号或选手账号'),
             is_active, session['admin_id']))
        cnt += 1
    _conn.commit(); _conn.close()
    return jsonify({'success': True, 'count': cnt, 'warnings': warnings})

@app.route('/api/competitions/template')
@admin_required
def competition_template():
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = '赛事导入模板'
    hdr_fill = PatternFill("solid", fgColor="E8F4FF")
    req_fill = PatternFill("solid", fgColor="FFE8E0")
    note_fill = PatternFill("solid", fgColor="FFF8E1")
    bold = Font(bold=True, name='微软雅黑', size=10)
    center = Alignment(horizontal='center', vertical='center')
    hdrs = ['赛事名称*','地点','开始时间','结束时间','欢迎语',
            '组别','赛事说明','云相册链接','负责人','查询字段','查询提示','是否上线',
            '子管理员手机号','子管理员权限']
    ws.append(hdrs)
    for i, cell in enumerate(ws[1]):
        cell.fill = req_fill if '*' in hdrs[i] else hdr_fill
        cell.font = bold; cell.alignment = center
    ws.append(['2025 Code The Future全国大赛', '北京·国家会议中心', '2025-06-15 08:30',
               '2025-06-16 17:00', '欢迎参加比赛',
               '初级组,中级组,高级组', '请凭报名编号完成签到',
               'https://album.example.com/', '张老师', 'player_no,account',
               '请输入报名编号或账号', '1',
               '13800000001,13800000002', 'view,edit'])
    # 说明行
    note_row = ['子管理员手机号：逗号分隔多个手机号', '', '', '', '', '', '', '', '', '', '', '',
                '多个手机号逗号分隔', '对应权限：view=仅查看，edit=可编辑，数量需与手机号一一对应']
    ws.append(note_row)
    for cell in ws[3]:
        if cell.value:
            cell.fill = note_fill
            cell.font = Font(name='微软雅黑', size=9, italic=True, color='856404')
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22
    ws.row_dimensions[1].height = 22
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return send_file(out,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name='赛事导入模板.xlsx')


@app.route('/api/competitions/<int:cid>/extra_fields', methods=['GET'])
@admin_required
def get_extra_fields(cid):
    a = get_me()
    if not a: return jsonify({'error': '未授权'}), 401
    if not can_view_comp(a, cid): return jsonify({'error': '无权限'}), 403
    r = db_query("SELECT extra_fields FROM competitions WHERE id=?", (cid,), one=True)
    if not r: return jsonify({'error': '赛事不存在'}), 404
    return jsonify(json.loads(r['extra_fields'] or '[]'))

@app.route('/api/competitions/<int:cid>/extra_fields', methods=['PUT'])
@admin_required
def update_extra_fields(cid):
    """保存赛事的自定义字段配置
    body: [{"key":"ef_xxx","label":"字段名","show_checkin":true,"show_list":true}]
    """
    a = get_me()
    if not a: return jsonify({'error': '未授权'}), 401
    if not can_edit_comp(a, cid): return jsonify({'error': '无权限，需要编辑权限'}), 403
    fields = request.json or []
    # 校验并规范化，key 统一用 ef_ 前缀防止冲突
    safe = []
    seen_labels = set()
    for f in fields:
        label = str(f.get('label','')).strip()
        if not label or label in seen_labels: continue
        seen_labels.add(label)
        # 保持已有 key，新字段生成 key
        key = f.get('key','')
        if not key or not key.startswith('ef_'):
            key = 'ef_' + hashlib.md5(label.encode()).hexdigest()[:8]
        safe.append({
            'key': key,
            'label': label,
            'show_checkin': bool(f.get('show_checkin', False)),
            'show_list': bool(f.get('show_list', True)),
        })
    db_exec("UPDATE competitions SET extra_fields=? WHERE id=?",
            (json.dumps(safe, ensure_ascii=False), cid))
    return jsonify({'success': True, 'fields': safe})

# ═══════════════════════════════════════════════════════════════════
# PLAYERS
# ═══════════════════════════════════════════════════════════════════

@app.route('/api/players', methods=['GET'])
@admin_required
def list_players():
    cid = request.args.get('competition_id')
    if not cid: return jsonify([])
    a = get_me()
    if not a: return jsonify({'error': '未授权'}), 401
    if not can_view_comp(a, int(cid)): return jsonify({'error': '无权限'}), 403
    q = "SELECT * FROM players WHERE competition_id=?"
    params = [cid]
    for f, col in [('group', 'group_name'), ('date', 'comp_date'),
                   ('session', 'session'), ('shirt', 'shirt_size'),
                   ('school', 'school'), ('grade', 'grade')]:
        v = request.args.get(f, '')
        if v: q += f" AND {col}=?"; params.append(v)
    checked = request.args.get('checked', '')
    if checked != '': q += " AND checked_in=?"; params.append(int(checked))
    search = request.args.get('search', '').strip()
    if search:
        q += " AND (name LIKE ? OR player_no LIKE ? OR account LIKE ? OR school LIKE ?)"
        params.extend([f'%{search}%'] * 4)
    q += " ORDER BY id"
    rows = db_query(q, params)
    return jsonify([dict(r) for r in rows])

@app.route('/api/players', methods=['POST'])
@admin_required
def create_player():
    d = request.json or {}
    if not d.get('name'): return jsonify({'error': '姓名必填'}), 400
    a = get_me()
    if not a: return jsonify({'error': '未授权'}), 401
    if not can_edit_comp(a, int(d.get('competition_id', 0))): return jsonify({'error': '无权限，需要编辑权限'}), 403
    db_exec("""INSERT INTO players(competition_id,player_no,account,name,school,grade,
                    group_name,comp_date,session,seat_no,shirt_size,remark,package,extra_data) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                 (d['competition_id'], d.get('player_no',''), d.get('account',''), d['name'],
                  d.get('school',''), d.get('grade',''), d.get('group_name',''),
                  d.get('comp_date',''), d.get('session',''), d.get('seat_no',''),
                  d.get('shirt_size',''), d.get('remark',''),
                  d.get('package',''), json.dumps(d.get('extra_data',{}), ensure_ascii=False)))
    return jsonify({'success': True})

@app.route('/api/players/<int:pid>', methods=['PUT'])
@admin_required
def update_player(pid):
    d = request.json or {}
    a = get_me()
    if not a: return jsonify({'error': '未授权'}), 401
    p = db_query("SELECT competition_id FROM players WHERE id=?", (pid,), one=True)
    if not p or not can_edit_comp(a, p['competition_id']):
        return jsonify({'error': '无权限，需要编辑权限'}), 403
    # extra_data：合并新值到旧值（保留未修改的字段）
    orig = db_query("SELECT extra_data FROM players WHERE id=?", (pid,), one=True)
    try:
        old_extra = json.loads(orig['extra_data'] or '{}') if orig else {}
    except Exception:
        old_extra = {}
    if 'extra_data' in d and isinstance(d['extra_data'], dict):
        old_extra.update(d['extra_data'])
    extra_data_val = json.dumps(old_extra, ensure_ascii=False)
    db_exec("""UPDATE players SET player_no=?,account=?,name=?,school=?,grade=?,
                    group_name=?,comp_date=?,session=?,seat_no=?,shirt_size=?,
                    checked_in=?,checkin_time=?,remark=?,package=?,extra_data=? WHERE id=?""",
                 (d.get('player_no',''), d.get('account',''), d.get('name',''),
                  d.get('school',''), d.get('grade',''), d.get('group_name',''),
                  d.get('comp_date',''), d.get('session',''), d.get('seat_no',''),
                  d.get('shirt_size',''), d.get('checked_in',0),
                  d.get('checkin_time',''), d.get('remark',''),
                  d.get('package',''), extra_data_val, pid))
    return jsonify({'success': True})

@app.route('/api/players/<int:pid>', methods=['DELETE'])
@admin_required
def delete_player(pid):
    a = get_me()
    if not a: return jsonify({'error': '未授权'}), 401
    p = db_query("SELECT competition_id FROM players WHERE id=?", (pid,), one=True)
    if not p or not can_edit_comp(a, p['competition_id']):
        return jsonify({'error': '无权限，需要编辑权限'}), 403
    db_exec("DELETE FROM players WHERE id=?", (pid,))
    return jsonify({'success': True})

@app.route('/api/players/import', methods=['POST'])
@admin_required
def import_players():
    a = get_me()
    if not a: return jsonify({'error': '未授权'}), 401
    if not can(a, 'import_players'): return jsonify({'error': '无权限'}), 403
    cid = request.form.get('competition_id')
    if cid and not can_edit_comp(a, int(cid)): return jsonify({'error': '无权限，需要编辑权限'}), 403
    upload = request.files.get('file')
    if not cid or not upload: return jsonify({'error': '参数缺失'}), 400

    # 取赛事信息：合法组别 + 自定义字段
    comp_row = db_query("SELECT `groups`, extra_fields FROM competitions WHERE id=?", (cid,), one=True)
    if not comp_row: return jsonify({'error': '赛事不存在'}), 404
    valid_groups = json.loads(comp_row['groups'] or '[]')
    # 自定义字段名列表（label→key映射）
    extra_field_defs = json.loads(comp_row['extra_fields'] or '[]')
    # extra_field_defs 格式: [{"key":"ef_xxx","label":"字段名","show_checkin":true,"show_list":true}]
    extra_label_map = {ef['label']: ef['key'] for ef in extra_field_defs}

    # 已存在的 player_no / account
    exist_nos  = set(r['player_no'] for r in db_query(
        "SELECT player_no FROM players WHERE competition_id=? AND player_no!=''", (cid,)))
    exist_accs = set(r['account'] for r in db_query(
        "SELECT account FROM players WHERE competition_id=? AND account!=''", (cid,)))

    wb = openpyxl.load_workbook(upload, data_only=True)
    ws = wb.active
    hdrs = [str(c.value).strip().rstrip('*').strip() if c.value else '' for c in ws[1]]

    # 固定字段映射
    fm = {'报名编号':'player_no','账号':'account','姓名':'name','学校':'school',
          '年级':'grade','组别':'group_name','比赛日期':'comp_date','场次':'session',
          '座位号':'seat_no','衣服尺码':'shirt_size','备注':'remark','赛事包':'package'}

    skipped = []   # 跳过的行及原因
    to_insert = []
    file_nos  = set()
    file_accs = set()

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row): continue
        pdata = {}
        extra = {}
        for i, hdr_name in enumerate(hdrs):
            val = str(row[i]).strip() if i < len(row) and row[i] is not None else ''
            if hdr_name in fm:
                pdata[fm[hdr_name]] = val
            elif hdr_name in extra_label_map:
                extra[extra_label_map[hdr_name]] = val
        if not pdata.get('name'): continue
        pdata['extra_data'] = extra

        row_errors = []
        pno  = pdata.get('player_no', '')
        pacc = pdata.get('account', '')
        grp  = pdata.get('group_name', '')

        # 重复检测：重复则跳过该行
        if pno:
            if pno in exist_nos or pno in file_nos:
                row_errors.append(f'报名编号"{pno}"重复')
            else:
                file_nos.add(pno)
        if pacc:
            if pacc in exist_accs or pacc in file_accs:
                row_errors.append(f'账号"{pacc}"重复')
            else:
                file_accs.add(pacc)

        # 组别校验：不符则跳过该行
        if valid_groups and grp and grp not in valid_groups:
            row_errors.append(f'组别"{grp}"不在赛事组别中（可选：{"、".join(valid_groups)}）')

        if row_errors:
            skipped.append(f'第{row_idx}行「{pdata.get("name","")}」：{"；".join(row_errors)}，已跳过')
        else:
            to_insert.append(pdata)

    # 批量插入正常行
    if to_insert:
        sql = """INSERT INTO players(competition_id,player_no,account,name,school,grade,
                        group_name,comp_date,session,seat_no,shirt_size,remark,package,extra_data)
                        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)"""
        ops = [(sql, (cid, p.get('player_no',''), p.get('account',''), p.get('name',''),
                      p.get('school',''), p.get('grade',''), p.get('group_name',''),
                      p.get('comp_date',''), p.get('session',''), p.get('seat_no',''),
                      p.get('shirt_size',''), p.get('remark',''),
                      p.get('package',''), json.dumps(p.get('extra_data',{}), ensure_ascii=False)))
               for p in to_insert]
        db_exec_many(ops)

    return jsonify({
        'success': True,
        'count': len(to_insert),
        'skipped': len(skipped),
        'warnings': skipped  # 跳过的行详情
    })

# ── 功能3：批量删除 ──────────────────────────────────────────────
@app.route('/api/players/batch_delete', methods=['POST'])
@admin_required
def batch_delete_players():
    d = request.json or {}
    ids = d.get('ids', [])
    if not ids: return jsonify({'error': '未选择选手'}), 400
    a = get_me()
    if not a: return jsonify({'error': '未授权'}), 401
    if not a['is_main']:
        ph2 = ','.join('?' * len(ids))
        player_rows = db_query(f"SELECT DISTINCT competition_id FROM players WHERE id IN ({ph2})", ids)
        for pr in player_rows:
            if not can_edit_comp(a, pr['competition_id']):
                return jsonify({'error': '无权限，需要编辑权限'}), 403
    placeholders = ','.join('?' * len(ids))
    db_exec(f"DELETE FROM players WHERE id IN ({placeholders})", ids)
    return jsonify({'success': True, 'count': len(ids)})

# ── 功能3：批量修改 ──────────────────────────────────────────────
@app.route('/api/players/batch_update', methods=['POST'])
@admin_required
def batch_update_players():
    d = request.json or {}
    ids    = d.get('ids', [])
    fields = d.get('fields', {})   # e.g. {"group_name":"高级组","session":"下午场"}
    if not ids:   return jsonify({'error': '未选择选手'}), 400
    if not fields: return jsonify({'error': '未指定修改字段'}), 400
    allowed = {'player_no','account','name','school','grade','group_name',
               'comp_date','session','seat_no','shirt_size','remark','checked_in',
               'checkin_time','package','extra_data'}
    safe = {k: v for k, v in fields.items() if k in allowed}
    if not safe: return jsonify({'error': '没有合法字段'}), 400
    a = get_me()
    if not a: return jsonify({'error': '未授权'}), 401
    if not a['is_main']:
        ph2 = ','.join('?' * len(ids))
        player_rows = db_query(f"SELECT DISTINCT competition_id FROM players WHERE id IN ({ph2})", ids)
        for pr in player_rows:
            if not can_edit_comp(a, pr['competition_id']):
                return jsonify({'error': '无权限，需要编辑权限'}), 403
    # extra_data 需要逐行合并（不整体覆盖，保留其他自定义字段的值）
    new_extra = None
    if 'extra_data' in safe:
        try:
            new_extra = json.loads(safe.pop('extra_data'))
        except Exception:
            safe.pop('extra_data', None)
    placeholders = ','.join('?' * len(ids))
    if safe:
        set_clause = ', '.join(f"{k}=?" for k in safe)
        vals = list(safe.values())
        db_exec(f"UPDATE players SET {set_clause} WHERE id IN ({placeholders})", vals + ids)
    if new_extra:
        # 逐行读取旧 extra_data，合并新值后写回
        rows = db_query(f"SELECT id, extra_data FROM players WHERE id IN ({placeholders})", ids)
        ops = []
        for row in rows:
            try:
                old = json.loads(row['extra_data'] or '{}')
            except Exception:
                old = {}
            old.update(new_extra)
            ops.append(("UPDATE players SET extra_data=? WHERE id=?",
                        (json.dumps(old, ensure_ascii=False), row['id'])))
        if ops: db_exec_many(ops)
    return jsonify({'success': True, 'count': len(ids)})

# ── 功能3：赛事批量删除 ──────────────────────────────────────────
@app.route('/api/competitions/batch_delete', methods=['POST'])
@admin_required
def batch_delete_competitions():
    a = get_me()
    if not a: return jsonify({'error': '未授权'}), 401
    d = request.json or {}
    ids = d.get('ids', [])
    if not ids: return jsonify({'error': '未选择赛事'}), 400
    for cid in ids:
        if not admin_owns_comp(a, cid):
            return jsonify({'error': '无权限'}), 403
    placeholders = ','.join(['%s'] * len(ids))
    db_exec_many([
        (f"DELETE FROM checkin_logs WHERE competition_id IN ({placeholders})", ids),
        (f"DELETE FROM players WHERE competition_id IN ({placeholders})", ids),
        (f"DELETE FROM competitions WHERE id IN ({placeholders})", ids),
    ])
    return jsonify({'success': True})

# ── 功能3：带筛选的选手导出 ──────────────────────────────────────
@app.route('/api/players/export/<int:cid>')
@admin_required
def export_players(cid):
    a = get_me()
    if not a: return jsonify({'error': '未授权'}), 401
    if not can_view_comp(a, cid): return jsonify({'error': '无权限'}), 403
    # 支持筛选参数导出（功能3）
    q = "SELECT * FROM players WHERE competition_id=?"
    params = [cid]
    for f, col in [('group', 'group_name'), ('date', 'comp_date'),
                   ('session', 'session'), ('shirt', 'shirt_size'), ('school', 'school'), ('grade', 'grade')]:
        v = request.args.get(f, '')
        if v: q += f" AND {col}=?"; params.append(v)
    checked = request.args.get('checked', '')
    if checked != '': q += " AND checked_in=?"; params.append(int(checked))
    search = request.args.get('search', '').strip()
    if search:
        q += " AND (name LIKE ? OR player_no LIKE ? OR account LIKE ? OR school LIKE ?)"
        params.extend([f'%{search}%'] * 4)
    q += " ORDER BY id"
    players = db_query(q, params)
    comp = db_query("SELECT name FROM competitions WHERE id=?", (cid,), one=True)
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = '选手信息'
    hfill = PatternFill("solid", fgColor="050d1f")
    afill = PatternFill("solid", fgColor="E8F8FF")
    thin = Border(*[Side(style='thin', color='CBD5E1')] * 4)
    _ef_row = db_query('SELECT extra_fields FROM competitions WHERE id=?',(cid,),one=True)
    extra_field_defs = json.loads((_ef_row.get('extra_fields') if _ef_row else None) or '[]')
    extra_keys = [ef['key'] for ef in extra_field_defs]
    extra_labels = [ef['label'] for ef in extra_field_defs]
    hdrs = ['ID','报名编号','账号','姓名','学校','年级','组别','比赛日期','场次','座位号','衣服尺码','赛事包','是否签到','签到时间','签到人','签到人手机','备注'] + extra_labels
    ws.append(hdrs)
    for cell in ws[1]:
        cell.fill = hfill
        cell.font = Font(color="00d4ff", bold=True, name='微软雅黑', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin
    ws.row_dimensions[1].height = 24
    for ri, p in enumerate(players):
        ed = json.loads(p.get('extra_data') or '{}')
        row = [p['id'], p['player_no'], p['account'], p['name'], p['school'],
               p['grade'], p['group_name'], p['comp_date'], p['session'],
               p['seat_no'], p['shirt_size'], p.get('package',''),
               '✅ 已签到' if p['checked_in'] else '⏳ 未签到',
               p['checkin_time'] or '', p.get('checkin_operator_name','') or '',
               p.get('checkin_operator_phone','') or '', p['remark'] or ''] + [ed.get(k,'') for k in extra_keys]
        ws.append(row)
        fill = afill if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for cell in ws[ri + 2]:
            cell.fill = fill
            cell.font = Font(name='微软雅黑', size=10)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = thin
        ws.row_dimensions[ri + 2].height = 20
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(
            len(str(c.value or '')) * 2 + 4 for c in col)
    ws.freeze_panes = 'A2'
    out = io.BytesIO(); wb.save(out); out.seek(0)
    name = f"{comp['name']}_选手信息.xlsx" if comp else "选手信息.xlsx"
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=name)

@app.route('/api/players/template')
@admin_required
def player_template():
    a = get_me()
    if not a: return jsonify({'error': '未授权'}), 401
    # 支持按赛事 ID 生成含自定义字段的模板
    cid = request.args.get('cid')
    extra_field_defs = []
    comp_name = ''
    if cid:
        comp_row = db_query("SELECT name, extra_fields FROM competitions WHERE id=?", (cid,), one=True)
        if comp_row:
            comp_name = comp_row['name']
            extra_field_defs = json.loads(comp_row['extra_fields'] or '[]')
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = '导入模板'
    hfill = PatternFill("solid", fgColor="E8F8FF")
    req_fill = PatternFill("solid", fgColor="FFE8E0")
    extra_fill = PatternFill("solid", fgColor="F0FFF0")
    bold = Font(bold=True, name='微软雅黑')
    center = Alignment(horizontal='center')
    fixed_hdrs = ['报名编号','账号','姓名*','学校','年级','组别','比赛日期','场次','座位号','衣服尺码','备注','赛事包']
    extra_hdrs = [ef['label'] for ef in extra_field_defs]
    hdrs = fixed_hdrs + extra_hdrs
    ws.append(hdrs)
    for i, cell in enumerate(ws[1]):
        if '*' in hdrs[i]:
            cell.fill = req_fill
        elif i >= len(fixed_hdrs):
            cell.fill = extra_fill  # 自定义字段用绿色背景区分
        else:
            cell.fill = hfill
        cell.font = bold; cell.alignment = center
    # 示例行
    demo = ['IC001','user001','张三','北京实验小学','六年级','初级组','2025-06-15','上午场','A-001','M','','有']
    demo += ['' for _ in extra_hdrs]
    ws.append(demo)
    # 自定义字段说明行
    if extra_hdrs:
        note_fill = PatternFill("solid", fgColor="FFF8E1")
        note = [''] * len(fixed_hdrs) + [f'自定义字段：{l}' for l in extra_hdrs]
        ws.append(note)
        for i, cell in enumerate(ws[3]):
            if i >= len(fixed_hdrs) and cell.value:
                cell.fill = note_fill
                cell.font = Font(name='微软雅黑', size=9, italic=True, color='856404')
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18
    out = io.BytesIO(); wb.save(out); out.seek(0)
    fname = f'{comp_name}_导入模板.xlsx' if comp_name else '选手导入模板.xlsx'
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)

# ═══════════════════════════════════════════════════════════════════
# STATISTICS
# ═══════════════════════════════════════════════════════════════════

# ── 功能4：赛事地点列表（供统计筛选用）────────────────────────
@app.route('/api/competitions/locations')
@admin_required
def competition_locations():
    a = get_me()
    if not a: return jsonify({'error': '未授权'}), 401
    if a['is_main']:
        rows = db_query("SELECT DISTINCT location FROM competitions WHERE location!='' ORDER BY location")
        return jsonify([r['location'] for r in rows])
    else:
        all_comps = db_query("SELECT id,location,comp_admins,created_by FROM competitions WHERE location!=''")
        locs = sorted(set(c['location'] for c in all_comps if can_view_comp(a, c['id'])))
        return jsonify(locs)

@app.route('/api/stats/<int:cid>')
@admin_required
def stats(cid):
    a = get_me()
    if not a: return jsonify({'error': '未授权'}), 401
    if not can_view_comp(a, cid): return jsonify({'error': '无权限'}), 403
    # 功能4：支持按地点筛选（location 筛选的是赛事列表层，这里支持跨赛事按地点汇总）
    location = request.args.get('location', '').strip()
    if location:
        # 找出该地点下当前管理员有权限的所有赛事id
        if a['is_main']:
            cid_rows = db_query("SELECT id FROM competitions WHERE location=?", (location,))
        else:
            all_loc = db_query("SELECT id,comp_admins,created_by FROM competitions WHERE location=?",
                                   (location,))
            cid_rows = [r for r in all_loc if can_view_comp(a, r['id'])]
        cids = [r['id'] for r in cid_rows]
        if not cids:
                return jsonify({'total':0,'checked':0,'unchecked':0,'by_session':[],'by_group':[],
                            'by_shirt':[],'by_date':[],'recent':[],'location':location,'comp_names':[]})
        ph = ','.join(['%s']*len(cids))
        where = f"competition_id IN ({ph})"
        p = cids
        comp_names = [r['name'] for r in db_query(f"SELECT name FROM competitions WHERE id IN ({ph})", cids)]
    else:
        where = "competition_id=?"
        p = [cid]
        comp_names = []

    total = list(db_query(f"SELECT COUNT(*) FROM players WHERE {where}", p, one=True).values())[0]
    chk   = list(db_query(f"SELECT COUNT(*) FROM players WHERE {where} AND checked_in=1", p, one=True).values())[0]
    by_session = db_query(f"""SELECT comp_date,session,COUNT(*) total,CAST(SUM(checked_in) AS UNSIGNED) checked
        FROM players WHERE {where} GROUP BY comp_date,session ORDER BY comp_date,session""", p)
    by_group = db_query(f"""SELECT group_name,COUNT(*) total,CAST(SUM(checked_in) AS UNSIGNED) checked
        FROM players WHERE {where} GROUP BY group_name ORDER BY total DESC""", p)
    by_shirt = db_query(f"""SELECT shirt_size,COUNT(*) total FROM players
        WHERE {where} GROUP BY shirt_size ORDER BY total DESC""", p)
    by_date  = db_query(f"""SELECT comp_date,COUNT(*) total,CAST(SUM(checked_in) AS UNSIGNED) checked
        FROM players WHERE {where} GROUP BY comp_date ORDER BY comp_date""", p)
    recent   = db_query(f"""SELECT pl.name,pl.group_name,pl.session,l.checkin_time,
        l.operator_name,l.operator_phone
        FROM checkin_logs l JOIN players pl ON l.player_id=pl.id
        WHERE l.{where} ORDER BY l.checkin_time DESC LIMIT 10""", p)
    return jsonify({
        'total': total, 'checked': chk, 'unchecked': total - chk,
        'by_session': [dict(r) for r in by_session],
        'by_group':   [dict(r) for r in by_group],
        'by_shirt':   [dict(r) for r in by_shirt],
        'by_date':    [dict(r) for r in by_date],
        'recent':     [dict(r) for r in recent],
        'location':   location,
        'comp_names': comp_names,
    })

@app.route('/api/stats/export/<int:cid>')
@admin_required
def export_stats(cid):
    a = get_me()
    if not a: return jsonify({'error': '未授权'}), 401
    if not can_view_comp(a, cid): return jsonify({'error': '无权限'}), 403
    comp = db_query("SELECT name FROM competitions WHERE id=?", (cid,), one=True)
    total = list(db_query("SELECT COUNT(*) FROM players WHERE competition_id=?", (cid,), one=True).values())[0]
    chk = list(db_query("SELECT COUNT(*) FROM players WHERE competition_id=? AND checked_in=1", (cid,), one=True).values())[0]
    by_session = db_query("""SELECT comp_date,session,COUNT(*) total,CAST(SUM(checked_in) AS UNSIGNED) checked
        FROM players WHERE competition_id=? GROUP BY comp_date,session""", (cid,))
    by_group = db_query("""SELECT group_name,COUNT(*) total,CAST(SUM(checked_in) AS UNSIGNED) checked
        FROM players WHERE competition_id=? GROUP BY group_name""", (cid,))
    by_shirt = db_query("""SELECT shirt_size,COUNT(*) total FROM players WHERE competition_id=?
        GROUP BY shirt_size""", (cid,))
    wb = openpyxl.Workbook()
    hfill = PatternFill("solid", fgColor="050d1f")
    def add_sheet(title, headers, rows, ws=None):
        if ws is None: ws = wb.create_sheet(title)
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = hfill
            cell.font = Font(color="00d4ff", bold=True, name='微软雅黑')
            cell.alignment = Alignment(horizontal='center')
        for row in rows: ws.append(row)
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18
        return ws
    ws0 = wb.active; ws0.title = '总览'
    add_sheet('总览', ['指标','数值'],
              [['报名总人数',total],['已签到',chk],['未签到',total-chk],
               ['签到率',f"{round(chk/total*100,1) if total else 0}%"]], ws0)
    add_sheet('按场次', ['日期','场次','总人数','已签到','未签到'],
              [[r['comp_date'],r['session'],r['total'],r['checked'],r['total']-r['checked']]
               for r in by_session])
    add_sheet('按组别', ['组别','总人数','已签到','未签到'],
              [[r['group_name'],r['total'],r['checked'],r['total']-r['checked']]
               for r in by_group])
    add_sheet('衣服尺码', ['尺码','人数'],
              [[r['shirt_size'],r['total']] for r in by_shirt])
    out = io.BytesIO(); wb.save(out); out.seek(0)
    name = f"{comp['name']}_统计数据.xlsx" if comp else "统计数据.xlsx"
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=name)


# ═══════════════════════════════════════════════════════════════════
# ADMINS
# ═══════════════════════════════════════════════════════════════════


@app.route('/api/admins/template')
@admin_required
def admin_template():
    a = get_me()
    if not a: return jsonify({'error': '未授权'}), 401
    if not a['is_main']: return jsonify({'error': '无权限'}), 403
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = '管理员导入模板'
    hdr_fill = PatternFill("solid", fgColor="E8F4FF")
    req_fill = PatternFill("solid", fgColor="FFE8E0")
    bold = Font(bold=True, name='微软雅黑', size=10)
    center = Alignment(horizontal='center', vertical='center')
    hdrs = ['姓名*','手机号*','初始密码*','角色（可选，自由填写如：赛事负责人）','新增赛事','导入选手','查看统计','人员管理']
    ws.append(hdrs)
    for i, cell in enumerate(ws[1]):
        cell.fill = req_fill if '*' in hdrs[i] else hdr_fill
        cell.font = bold; cell.alignment = center
    ws.append(['张老师','13800000001','abc123','赛事负责人','','','',''])
    ws.append(['李老师','13800000002','abc123','','否','否','是','否'])
    note_fill = PatternFill("solid", fgColor="FFF8E1")
    note = ['权限列填"是"或"1"表示开启，其余表示关闭','','','新增赛事','导入选手','查看统计','人员管理']
    ws.append(note)
    for cell in ws[3]:
        if cell.value:
            cell.fill = note_fill
            cell.font = Font(name='微软雅黑', size=9, italic=True, color='856404')
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return send_file(out,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name='管理员导入模板.xlsx')

@app.route('/api/admins/import', methods=['POST'])
@admin_required
def import_admins():
    a = get_me()
    if not a: return jsonify({'error': '未授权'}), 401
    if not a['is_main']: return jsonify({'error': '无权限'}), 403
    upload = request.files.get('file')
    if not upload: return jsonify({'error': '请上传文件'}), 400
    wb = openpyxl.load_workbook(upload, data_only=True)
    ws = wb.active
    hdrs = [str(c.value).strip().rstrip('*').strip() if c.value else '' for c in ws[1]]
    perm_map = {'新增赛事':'add_competition','导入选手':'import_players',
                '查看统计':'checkin_stats','人员管理':'manage_admins'}
    _conn = db(); _cur = _conn.cursor(); cnt = 0; warnings = []
    for row_i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row): continue
        d = dict(zip(hdrs, [str(v).strip() if v is not None else '' for v in row]))
        name = d.get('姓名','').strip()
        phone = d.get('手机号','').strip()
        pwd = d.get('初始密码','').strip()
        if not name or not phone or not pwd:
            warnings.append(f'第{row_i}行：姓名/手机号/密码不能为空，已跳过'); continue
        if len(pwd) < 6:
            warnings.append(f'第{row_i}行：{name} 密码少于6位，已跳过'); continue
        perms = {}
        for col, key in perm_map.items():
            val = d.get(col, '')
            perms[key] = val in ('是','1','yes','true','TRUE','YES')
        role_nm = d.get('角色','').strip()
        try:
            _cur.execute(
                "INSERT INTO admins(name,phone,password,role_name,permissions) VALUES(%s,%s,%s,%s,%s)",
                (name, phone, sha(pwd), role_nm, json.dumps(perms, ensure_ascii=False)))
            cnt += 1
        except Exception:
            _conn.rollback()
            warnings.append(f'第{row_i}行：手机号 {phone} 已存在，已跳过')
    _conn.commit(); _conn.close()
    return jsonify({'success': True, 'count': cnt, 'warnings': warnings})

@app.route('/api/admins/batch_delete', methods=['POST'])
@admin_required
def batch_delete_admins():
    a = get_me()
    if not a: return jsonify({'error': '未授权'}), 401
    if not a['is_main']: return jsonify({'error': '无权限'}), 403
    ids = (request.json or {}).get('ids', [])
    if not ids: return jsonify({'error': '未选择管理员'}), 400
    ph = ','.join(['%s']*len(ids))
    db_exec(f"DELETE FROM admins WHERE id IN ({ph}) AND is_main=0", ids)
    return jsonify({'success': True})


@app.route('/api/admins/batch_update', methods=['POST'])
@admin_required
def batch_update_admins():
    a = get_me()
    if not a: return jsonify({'error': '未授权'}), 401
    if not a['is_main']: return jsonify({'error': '无权限'}), 403
    d = request.json or {}
    ids = d.get('ids', [])
    if not ids: return jsonify({'error': '未选择管理员'}), 400
    ph = ','.join(['%s']*len(ids))
    if 'role_name' in d:
        db_exec(f"UPDATE admins SET role_name=? WHERE id IN ({ph}) AND is_main=0",
                     [d['role_name']] + ids)
    if 'permissions' in d:
        db_exec(f"UPDATE admins SET permissions=? WHERE id IN ({ph}) AND is_main=0",
                     [json.dumps(d['permissions'], ensure_ascii=False)] + ids)
    if 'is_active' in d:
        db_exec(f"UPDATE admins SET is_active=? WHERE id IN ({ph}) AND is_main=0",
                     [1 if d['is_active'] else 0] + ids)
    if d.get('reset_password'):
        db_exec(f"UPDATE admins SET password=? WHERE id IN ({ph}) AND is_main=0",
                     [sha(d['reset_password'])] + ids)
    return jsonify({'success': True})

@app.route('/api/admins', methods=['GET'])
@admin_required
def list_admins():
    a = get_me()
    if not a: return jsonify({'error': '未授权'}), 401
    if not a['is_main']: return jsonify({'error': '无权限'}), 403
    rows = db_query("SELECT id,name,phone,is_main,role_name,is_active,permissions,created_at FROM admins ORDER BY id")
    return jsonify([dict(r) for r in rows])

@app.route('/api/admins', methods=['POST'])
@admin_required
def create_admin():
    a = get_me()
    if not a: return jsonify({'error': '未授权'}), 401
    if not a['is_main']: return jsonify({'error': '无权限'}), 403
    d = request.json or {}
    if not d.get('name') or not d.get('phone') or not d.get('password'):
        return jsonify({'error': '姓名、手机号、密码均为必填'}), 400
    if len(d.get('password','')) < 6:
        return jsonify({'error': '密码至少6位'}), 400
    try:
        db_exec("INSERT INTO admins(name,phone,password,role_name,is_active,permissions) VALUES(?,?,?,?,?,?)",
                     (d['name'], d['phone'], sha(d['password']),
                      d.get('role_name',''),
                      1 if d.get('is_active',1) else 0,
                      json.dumps(d.get('permissions', {}), ensure_ascii=False)))
    except Exception:
        return jsonify({'error': '手机号已存在'}), 400
    return jsonify({'success': True})

@app.route('/api/admins/<int:aid>', methods=['PUT'])
@admin_required
def update_admin(aid):
    a = get_me(); d = request.json or {}
    if not a: return jsonify({'error': '未授权'}), 401
    if not a['is_main'] and a['id'] != aid: return jsonify({'error': '无权限'}), 403
    if a['is_main']:
        db_exec("UPDATE admins SET name=?,role_name=?,is_active=?,permissions=? WHERE id=?",
                     (d.get('name',''), d.get('role_name',''),
                      1 if d.get('is_active',1) else 0,
                      json.dumps(d.get('permissions',{}), ensure_ascii=False), aid))
    if d.get('password'):
        db_exec("UPDATE admins SET password=? WHERE id=?", (sha(d['password']), aid))
    return jsonify({'success': True})

@app.route('/api/admins/<int:aid>', methods=['DELETE'])
@admin_required
def delete_admin(aid):
    a = get_me()
    if not a: return jsonify({'error': '未授权'}), 401
    if not a['is_main']: return jsonify({'error': '无权限'}), 403
    db_exec("DELETE FROM admins WHERE id=? AND is_main=0", (aid,))
    return jsonify({'success': True})

if __name__ == '__main__':
    init_db()
    print("\n" + "="*55)
    print("  🚀 ICode 签到管理系统 v3.4")
    print("  选手签到: http://localhost:5001/")
    print("  管理后台: http://localhost:5001/admin")
    print("  账号: admin  密码: admin123")
    print("="*55 + "\n")
    app.run(debug=False, host='0.0.0.0', port=5001)
